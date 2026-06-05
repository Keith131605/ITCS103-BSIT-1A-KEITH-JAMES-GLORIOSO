import os
import openpyxl as op
import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime  

DB_FILE = "Glorioso_DatabaseDB.xlsx"

if not os.path.exists(DB_FILE):
    wb = op.Workbook()
    ws = wb.active
    ws.append(["Booking ID", "Room Name", "Date (YYYY-MM-DD)", "Time Slot", "Host Name", "Purpose"])
    wb.save(DB_FILE)

def display_excel():
    table.unbind("<<TreeviewSelect>>")
    workbook = op.load_workbook(DB_FILE)
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            table.insert("", tk.END, values=row)
            
    table.bind("<<TreeviewSelect>>", auto_populate)

def validate_input():
    room = room_combo.get().strip() 
    date = date_entry.get().strip()
    time_slot = time_combo.get().strip()
    host = host_entry.get().strip()
    purpose = purpose_entry.get().strip()

    if not room or not date or not time_slot or not host or not purpose:
        messagebox.showerror("Error", "All fields are required.")
        return False

    try:
        datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Error", "Date must be in the valid YYYY-MM-DD format.")
        return False

    return True

def auto_populate(event):
    selected = table.focus()
    if not selected:
        return

    values = table.item(selected, "values")

    if values and len(values) >= 6:
        clear_entries()
        room_combo.set(values[1])
        date_entry.insert(0, values[2])
        time_combo.set(values[3])
        host_entry.insert(0, values[4])
        purpose_entry.insert(0, values[5])

def saving():
    if not validate_input(): 
        return
        
    room = room_combo.get().strip()
    date = date_entry.get().strip()
    time_slot = time_combo.get().strip()
    host = host_entry.get().strip()
    purpose = purpose_entry.get().strip()

    workbook = op.load_workbook(DB_FILE)
    sheet = workbook.active

    if sheet.max_row == 1:
        new_id = 1
    else:
        last_id_value = sheet.cell(row=sheet.max_row, column=1).value
        new_id = int(last_id_value) + 1 if last_id_value else sheet.max_row

    sheet.append([new_id, room, date, time_slot, host, purpose])
    workbook.save(DB_FILE)
    messagebox.showinfo("Success", "Booking reserved successfully!")
    clear_entries()
    display_excel()

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a booking from the table first.")
        return
    if not validate_input():  
        return

    values = table.item(selected, "values")
    if not values:
        return
        
    booking_id = values[0]

    room = room_combo.get().strip() 
    date = date_entry.get().strip()
    time_slot = time_combo.get().strip()
    host = host_entry.get().strip()
    purpose = purpose_entry.get().strip()

    workbook = op.load_workbook(DB_FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(booking_id):
            row[1].value = room
            row[2].value = date
            row[3].value = time_slot
            row[4].value = host
            row[5].value = purpose
            break

    workbook.save(DB_FILE)
    messagebox.showinfo("Success", "Booking updated successfully!")
    clear_entries()
    display_excel()

def delete_record():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Please select a booking to delete.")
        return
    
    values = table.item(selected, "values")
    if not values:
        return
        
    booking_id = values[0]

    confirm = messagebox.askyesno("Confirm", "Are you sure you want to cancel this booking?")
    if not confirm:
        return

    workbook = op.load_workbook(DB_FILE)
    sheet = workbook.active

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(booking_id):
            sheet.delete_rows(row_idx, 1)
            break 

    workbook.save(DB_FILE)
    messagebox.showinfo("Success", "Booking deleted successfully!")
    clear_entries()
    display_excel()

def clear_entries():
    room_combo.set('') 
    date_entry.delete(0, tk.END)
    time_combo.set('')
    host_entry.delete(0, tk.END)
    purpose_entry.delete(0, tk.END)

window = tk.Tk()
window.title("Meeting Room Booking System")
window.configure(bg="lightblue")

title = tk.Label(window, text="Meeting Room Booking System", font=("Arial", 16, "bold"), bg="lightblue", fg="black")
title.grid(row=0, column=0, columnspan=6, pady=10)

genframe = tk.Frame(window, bg="lightblue", bd=2, relief="groove")
genframe.grid(row=1, column=0, columnspan=6, padx=15, pady=10)

rooms_list = ["Conference Room A", "Conference Room B", "Boardroom", "meeting Room 1", "meeting Room 2", "Main Hall"]
room_combo = ttk.Combobox(genframe, values=rooms_list, font=("Poppins", 11), state="readonly")
room_combo.grid(row=2, column=1, columnspan=2, padx=10, pady=(10, 0))

room_label = tk.Label(genframe, text="Room Name / Number", font=("Poppins", 9, "bold"), bg="lightblue", fg="black")
room_label.grid(row=3, column=1, columnspan=2)

date_entry = tk.Entry(genframe, font=("Poppins", 11))
date_entry.grid(row=2, column=3, columnspan=2, padx=10, pady=(10, 0))

date_label = tk.Label(genframe, text="Date (YYYY-MM-DD)", font=("Poppins", 9, "bold"), bg="lightblue", fg="black")
date_label.grid(row=3, column=3, columnspan=2)

time_slots = ["08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM","11:00 AM - 12:00 PM","12:00 PM - 01:00 PM","01:00 PM - 02:00 PM","02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM"]
time_combo = ttk.Combobox(genframe, values=time_slots, font=("Poppins", 11), state="readonly")
time_combo.grid(row=4, column=1, columnspan=2, padx=10, pady=(10, 0))

time_label = tk.Label(genframe, text="Time Slot", font=("Poppins", 9, "bold"), bg="lightblue", fg="black")
time_label.grid(row=5, column=1, columnspan=2)

host_entry = tk.Entry(genframe, font=("Poppins", 11))
host_entry.grid(row=4, column=3, columnspan=2, padx=10, pady=(10, 0))

host_label = tk.Label(genframe, text="Host / Reserved By", font=("Poppins", 9, "bold"), bg="lightblue", fg="black")
host_label.grid(row=5, column=3, columnspan=2)

purpose_entry = tk.Entry(genframe, font=("Poppins", 11))
purpose_entry.grid(row=6, column=1, columnspan=4, padx=10, pady=(10, 0), sticky="ew")

purpose_label = tk.Label(genframe, text="Meeting Purpose", font=("Poppins", 9, "bold"), bg="lightblue", fg="black")
purpose_label.grid(row=7, column=1, columnspan=4)

submit_btn = tk.Button(window, text="Book Room", font=("Poppins", 10, "bold"), bg="white", fg="black", command=saving)
submit_btn.grid(row=8, column=1, pady=(15, 15), padx=5)

update_btn = tk.Button(window, text="Update Booking", font=("Poppins", 10, "bold"), bg="green", fg="white", command=update)
update_btn.grid(row=8, column=2, pady=(15, 15), padx=5)

delete_btn = tk.Button(window, text="Cancel Booking", bg="red", fg="white", font=("Poppins", 10, "bold"), command=delete_record)
delete_btn.grid(row=8, column=3, pady=(15, 15), padx=5)

clear_btn = tk.Button(window, text="Clear Fields", bg="gray", fg="white", font=("Poppins", 10, "bold"), command=clear_entries)
clear_btn.grid(row=8, column=4, pady=(15, 15), padx=5)

headers = ("Booking ID", "Room Name", "Date", "Time Slot", "Host Name", "Purpose")
table = ttk.Treeview(window, columns=headers, show="headings")

for heading in headers:
    table.heading(heading, text=heading)
    table.column(heading, width=130, anchor="center")
table.grid(row=9, column=0, columnspan=6, padx=15, pady=15)

display_excel()

window.mainloop()