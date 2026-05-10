import tkinter as tk
import openpyxl as op
from tkinter import ttk, messagebox

CURRENT_YEAR = 2026
FILE_NAME = "favorite_people.xlsx"

class FaveRecorder:
    def __init__(item, window):
        item.root = window
        item.root.title("Age Calculator & Fave Recorder")
        item.root.geometry("850x600")
        item.root.configure(bg="green")
        
        item.id_counter = 1

        tk.Label(window, text="RECORD FAVORITE PEOPLE", font=("Arial", 16, "bold"), 
                 bg="green", fg="white").pack(pady=10)

       
        input_frame = tk.Frame(window, bg="green")
        input_frame.pack(pady=10)

        label_theme = {"bg": "green", "fg": "white", "font": ("Arial", 10, "bold")}
        
   
        tk.Label(input_frame, text="First Name:").grid(row=0, column=0, padx=5, pady=5)
        item.entry_first = ttk.Entry(input_frame, width=25)
        item.entry_first.grid(row=0, column=1)

  
        tk.Label(input_frame, text="Middle Name:").grid(row=1, column=0, padx=5, pady=5)
        item.entry_middle = ttk.Entry(input_frame, width=25)
        item.entry_middle.grid(row=1, column=1)

   
        tk.Label(input_frame, text="Last Name:").grid(row=2, column=0, padx=5, pady=5)
        item.entry_last = ttk.Entry(input_frame, width=25)
        item.entry_last.grid(row=2, column=1)

       
        tk.Label(input_frame, text="Birth Year:").grid(row=3, column=0, padx=5, pady=5)
        item.entry_birth = ttk.Entry(input_frame, width=25)
        item.entry_birth.grid(row=3, column=1)

    
        ttk.Button(window, text="Add & Save", command=item.add_and_save).pack(pady=10)

   
        item.table = ttk.Treeview(window, columns=("ID", "FName", "MName", "LName", "BYear", "Age"), show='headings', height=10)
        item.table.pack(pady=20, padx=20, fill=tk.BOTH, expand=True)

        item.table.heading("ID", text="ID")
        item.table.column("ID", width=40, anchor="center")
        item.table.heading("FName", text="First Name")
        item.table.column("FName", width=130)
        item.table.heading("MName", text="Middle Name")
        item.table.column("MName", width=130)
        item.table.heading("LName", text="Last Name")
        item.table.column("LName", width=130)
        item.table.heading("BYear", text="Birth Year")
        item.table.column("BYear", width=80, anchor="center")
        item.table.heading("Age", text="Age")
        item.table.column("Age", width=60, anchor="center")

        
    def display(item):
        """Refreshes the table with data from the Excel file."""
        try:
            workbook = op.load_workbook(FILE_NAME)
            sheet = workbook.active

            for row in item.table.get_children():
                item.table.delete(row)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                item.table.insert("", tk.END, values=row)
                
        except FileNotFoundError:
            pass

    def add_and_save(item):
        if item.id_counter > 3:
            messagebox.showinfo("Limit Reached", "3 people have already been recorded.")
            return

        try:
            fname = item.entry_first.get().strip()
            mname = item.entry_middle.get().strip()
            lname = item.entry_last.get().strip()
            byear = int(item.entry_birth.get().strip())
            age = CURRENT_YEAR - byear

       
            if item.id_counter == 1:
                wb = op.Workbook()
                ws = wb.active
                ws.title = "Records"
               
                ws.append(["ID", "First Name", "Middle Name", "Last Name", "Birth Year", "Age"])
            else:
                wb = op.load_workbook(FILE_NAME)
                ws = wb.active

            
            ws.append([item.id_counter, fname, mname, lname, byear, age])
            wb.save(FILE_NAME)

            item.id_counter += 1
            
            item.entry_first.delete(0, tk.END)
            item.entry_middle.delete(0, tk.END)
            item.entry_last.delete(0, tk.END)
            item.entry_birth.delete(0, tk.END)

            item.display()

            if item.id_counter > 3:
                messagebox.showinfo("Success", "All records successfully saved!")

        except ValueError:
            messagebox.showerror("Error", "Please enter a valid Birth Year.")

if __name__ == "__main__":
    window = tk.Tk()
    app = FaveRecorder(window)
    window.mainloop()